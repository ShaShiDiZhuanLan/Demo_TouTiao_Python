# coding: utf-8
"""
Author: 沙振宇
Update Time: 2019-12-24
Info: 数据整理（先获取全部数据，再过滤）
    # 1、获取来源__今日头条为数据来源：https://www.toutiao.com/
    # 2、抓取内容__搜索表达式(时间为 2019年12月25)
    # 3、数据处理__然后先初步看下这样抓取下来的新闻内容是什么，条数是多少
    # 4、词频统计__对这些新闻内容进行分词，然后词频统计。
"""
import requests
import random
import json
from openpyxl import Workbook
import time
import hashlib
import os
import datetime

# https://www.toutiao.com
# 参考 https://www.toutiao.com/api/pc/feed/?max_behot_time=1577172421&category=__all__&utm_source=toutiao&widen=1&tadrequire=true&as=A195EE50617CA9A&cp=5E016CCA69CAAE1&_signature=BH146AAgEBP2UkqvuhP-rgR9ePAAFqC
class TouTiao:
    # 初始化
    def __init__(self):
        self.start_url = 'https://www.toutiao.com/api/pc/feed/?category=news_hot&utm_source=toutiao&widen=1&max_behot_time='
        self.url = 'https://www.toutiao.com'
        self.user_agent_list = [
            'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1464.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/31.0.1650.16 Safari/537.36',
            'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.3319.102 Safari/537.36',
            'Mozilla/5.0 (X11; CrOS i686 3912.101.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.116 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.93 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/32.0.1667.0 Safari/537.36',
            'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:17.0) Gecko/20100101 Firefox/17.0.6',
            'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1468.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2224.3 Safari/537.36',
            'Mozilla/5.0 (X11; CrOS i686 3912.101.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.116 Safari/537.36'
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36']  # 浏览器header
        self.user_agent = random.choice(self.user_agent_list)
        self.headers = {
            'user-agent': self.user_agent
        }
        self.cookies = {'tt_webid': '6773835217905321479'}  # 此处cookies可从浏览器中查找，为了避免被头条禁止爬虫
        self.max_behot_time = '0'  # 链接参数
        self.title = []  # 存储新闻标题
        self.source_url = []  # 存储新闻的链接
        self.total_url = []  # 存储新闻的完整链接
        self.source = []  # 存储发布新闻的公众号
        self.media_url = {}  # 存储公众号的完整链接

    # 该函数主要是为了获取as和cp参数 程序参考今日头条中的加密js文件：home_4abea46.js
    # r如果没有填cur_time就默认取当前时间
    def get_as_cp(self, cur_time = round(time.time())):
        # 1577175580 - 2019 / 12 / 24 16: 19:40
        # 1575129600 - 2019 / 12 / 01 00: 00:00
        # 1559318400 - 2019 / 06 / 01 00: 00:00
        now = cur_time
        # print('now:', now)  # 获取当前计算机时间
        time_local = time.localtime(now)
        dt = time.strftime("%Y-%m-%d %H:%M:%S", time_local)
        print('时间:', dt)  # 获取当前时间
        e = hex(int(now)).upper()[2:]  # hex()转换一个整数对象为16进制的字符串表示
        a = hashlib.md5()  # hashlib.md5().hexdigest()创建hash对象并返回16进制结果
        a.update(str(int(now)).encode('utf-8'))
        i = a.hexdigest().upper()
        if len(e) != 8:
            zz = {'as': '479BB4B7254C150',
                  'cp': '7E0AC8874BB0985'}
            return zz
        n = i[:5]
        a = i[-5:]
        r = ''
        s = ''
        for i in range(5):
            s = s + n[i] + e[i]
        for j in range(5):
            r = r + e[j + 3] + a[j]
        zz = {
            'as': 'A1' + s + e[-3:],
            'cp': e[0:3] + r + 'E1'
        }
        return zz

    # 解析网页函数
    def getdata(self, url, headers, cookies):
        r = requests.get(url, headers=headers, cookies=cookies)
        # print("url:",url)
        # print("r.text:",r.text)
        try:
            data = json.loads(r.text)
            return data
        except ValueError:
            # print("------- getdata ValueError --------------")
            return {}

    # 存储数据到文件
    def savedata(self, title, total_url, source, media_url):
        # 存储数据到xlxs文件
        wb = Workbook()
        if not os.path.isdir(os.getcwd() + '/result'):  # 判断文件夹是否存在
            os.makedirs(os.getcwd() + '/result')  # 新建存储文件夹
        filename = os.getcwd() + '/result/result-' + datetime.datetime.now().strftime(
            '%Y-%m-%d-%H-%m') + '.xlsx'  # 新建存储结果的excel文件
        ws = wb.active
        ws.title = 'data'  # 更改工作表的标题
        ws['A1'] = '标题'  # 对表格加入标题
        ws['B1'] = '新闻链接'
        ws['C1'] = '头条号'
        ws['D1'] = '头条号链接'
        for row in range(2, len(title) + 2):  # 将数据写入表格
            _ = ws.cell(column=1, row=row, value=title[row - 2])
            _ = ws.cell(column=2, row=row, value=total_url[row - 2])
            _ = ws.cell(column=3, row=row, value=source[row - 2])
            _ = ws.cell(column=4, row=row, value=media_url[source[row - 2]])

        wb.save(filename=filename)  # 保存文件

    def main(self, max_behot_time, title, source_url, total_url, source, media_url):  # 主函数
        # 此处的数字类似于你刷新新闻的次数，正常情况下刷新一次会出现10条新闻，但夜存在少于10条的情况；所以最后的结果并不一定是10的倍数
        time_2019_12_26_00_00_00 = 1577289600   # 2019/12/26 00: 00:00
        time_2019_12_25_00_00_00 = 1577203200   # 2019/12/25 00: 00:00
        cur_time = time_2019_12_25_00_00_00

        while cur_time < time_2019_12_26_00_00_00:
            cur_time = cur_time + 30 # 隔30秒过滤一下
            ascp = self.get_as_cp(cur_time)  # 获取as和cp参数的函数
            demo = self.getdata(
                self.start_url + max_behot_time + '&max_behot_time_tmp=' + max_behot_time + '&tadrequire=true&as=' + ascp['as'] + '&cp=' + ascp['cp'], self.headers, self.cookies)
            print("demo:",demo)
            if 'data' in demo:
                for j in range(len(demo['data'])):
                    title_tmp = demo['data'][j]['title']
                    # if '失业' in title_tmp:
                    time_local = time.localtime(cur_time)
                    dt = time.strftime("%Y-%m-%d %H:%M:%S", time_local)
                    print('-- 时间:%s %s'%(dt, title_tmp))  # 获取当前时间
                    # if ('人工智能' in title_tmp) or ('AI' in title_tmp) or ('ai' in title_tmp):
                    if demo['data'][j]['title'] not in title:
                        title.append(demo['data'][j]['title'])  # 获取新闻标题
                        source_url.append(demo['data'][j]['source_url'])  # 获取新闻链接
                        url_gzh = demo['data'][j]['source']
                        source.append(url_gzh)  # 获取发布新闻的公众号
                        print('title_data:', demo['data'][j]['title'])
                        print("url_gzh:",url_gzh)
                        if url_gzh not in media_url:
                            media_url[url_gzh] = self.url + demo['data'][j]['media_url']  # 获取公众号链接
                max_behot_time = str(demo['next']['max_behot_time'])  # 获取下一个链接的max_behot_time参数的值

        for index in range(len(title)):
            print('标题：', title[index])
            if 'https' not in source_url[index]:
                total_url.append(self.url + source_url[index])
                print('新闻链接：', self.url + source_url[index])
            else:
                print('新闻链接：', source_url[index])
                total_url.append(source_url[index])
                print('源链接：', self.url+source_url[index])
            print('头条号(新闻网站)：', source[index])
            # print('获取的新闻数量：', len(title))

if __name__ == "__main__":
    tt = TouTiao()
    tt.main(tt.max_behot_time, tt.title, tt.source_url, tt.total_url, tt.source, tt.media_url)
    print("终于获取完了，开始保存数据")
    tt.savedata(tt.title, tt.total_url, tt.source, tt.media_url)
    print("始保存数据成功")
